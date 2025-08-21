#--- coding=utf-8
#--- @Author: Yongsheng.Guo@ansys.com, Henry.he@ansys.com,Yang.zhao@ansys.com
#--- @Time: 20240830


'''
read and get config in excel
'''
import os
import sys
isIronpython = "IronPython" in sys.version

def XlsReader(path):
    ext = os.path.splitext(path)
    if ext[1].lower() == ".xlsx":
        return _XlsxReader(path)
    elif ext[1].lower() == ".xls":
        return _XlsReader(path)
    else:
        raise Exception("Unsupported file format,must be xlsx or xls.")


class _XlsxReader(object):
    def __init__(self, path):
        self.path = path
        self._workbook = None  # 缓存 workbook 对象

    def _load_workbook(self):
        """加载 Excel 文件（缓存机制）"""
        import openpyxl
        if not self._workbook:
            self._workbook = openpyxl.load_workbook(self.path, data_only=True)
        return self._workbook

    def _close_workbook(self):
        """关闭 workbook（释放资源）"""
        if self._workbook:
            self._workbook.close()
            self._workbook = None

    def _parse_value(self, value):
        """统一处理单元格值的类型转换"""
        if value is None:
            return ""
        elif isinstance(value, (str)):  # Python 2.7 需要处理 unicode
            lower_val = value.lower()
            if lower_val == "none":
                return ""
            elif lower_val == "true":
                return True
            elif lower_val == "false":
                return False
            else:
                return value.strip()  # 去除字符串两端空格
        else:
            return str(value)

    def readSheet(self, sheet, header_row=1):
        """
        读取单个工作表的数据，返回字典列表。
        Args:
            sheet: openpyxl 工作表对象
            header_row: 表头所在行（从 1 开始计数）
        Returns:
            List of dict, 例如: [{"Name": "Alice", "Age": 25}, ...]
        """
        headers = [cell.value for cell in sheet[header_row]]
        data = []

        for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            # 跳过空行（所有列均为 None）
            if all(cell is None for cell in row):
                continue
            row_dict = {
                headers[col]: self._parse_value(row[col])
                for col in range(len(headers)) if col < len(row)  # 防止列数不匹配
            }
            data.append(row_dict)

        return data

    def getSheetData(self, sheetName=None):
        """
        获取指定工作表的数据。
        Args:
            sheetName: 
                - str: 工作表名称（如 "Sheet1"）
                - int: 工作表索引（从 0 开始）
                - list: 多个工作表名称或索引
                - None: 返回活动工作表数据
        Returns:
            - 如果 sheetName 是 list，返回 {sheetName: data} 的字典
            - 否则返回单个工作表的数据列表
        """
        workbook = self._load_workbook()
        try:
            # 处理多工作表请求
            if isinstance(sheetName, list):
                result = {}
                for name in sheetName:
                    resolved_name = self._get_sheet_name(workbook, name)
                    result[resolved_name] = self.readSheet(workbook[resolved_name])
                return result
            # 处理单工作表请求
            elif sheetName is not None:
                resolved_name = self._get_sheet_name(workbook, sheetName)
                return self.readSheet(workbook[resolved_name])
            else:
                return self.readSheet(workbook.active)
        finally:
            self._close_workbook()

    def _get_sheet_name(self, workbook, sheet_ref):
        """根据名称或索引返回工作表名称"""
        if isinstance(sheet_ref, (str)):
            sheet_ref = sheet_ref.lower()
            for name in workbook.sheetnames:
                if name.lower() == sheet_ref:
                    return name
            raise ValueError("未找到工作表: {}".format(sheet_ref))
        elif isinstance(sheet_ref, int):
            if sheet_ref < 0 or sheet_ref >= len(workbook.sheetnames):
                raise IndexError("工作表索引越界: {}".format(sheet_ref))
            return workbook.sheetnames[sheet_ref]
        else:
            raise TypeError("sheet_name 必须是 str 或 int")

    def getAllSheetData(self):
        """获取所有工作表的数据，返回 {sheet_name: data} 的字典"""
        workbook = self._load_workbook()
        try:
            return {
                sheet_name: self.readSheet(workbook[sheet_name])
                for sheet_name in workbook.sheetnames
            }
        finally:
            self._close_workbook()

    def readAll(self):
        """兼容旧接口，等同于 getAllSheetData"""
        return self.getAllSheetData()

# -*- coding: utf-8 -*-


class _XlsReader(object):
    def __init__(self, path):
        self.path = path
        self._workbook = None  # 缓存 workbook 对象

    def _load_workbook(self):
        """加载 Excel 文件（缓存机制）"""
        import xlrd
        if not self._workbook:
            self._workbook = xlrd.open_workbook(self.path)
        return self._workbook

    def _close_workbook(self):
        """关闭 workbook（xlrd 不需要显式关闭，但保持接口一致）"""
        self._workbook = None

    def _parse_value(self, value):
        """统一处理单元格值的类型转换"""
        if isinstance(value, (str)):
            lower_val = value.lower()
            if lower_val == "none":
                return ""
            elif lower_val == "true":
                return True
            elif lower_val == "false":
                return False
            else:
                return value.strip()  # 去除字符串两端空格
        elif value is None or value == "":
            return ""
        else:
            return value

    def readSheet(self, sheet, header_row=0):
        """
        读取单个工作表的数据，返回字典列表。
        Args:
            sheet: xlrd 的 sheet 对象
            header_row: 表头所在行（从 0 开始计数，默认第 1 行）
        Returns:
            List of dict, 例如: [{"Name": "Alice", "Age": 25}, ...]
        """
        headers = sheet.row_values(header_row)
        data = []

        for row_idx in range(header_row + 1, sheet.nrows):
            row_values = sheet.row_values(row_idx)
            # 跳过空行（所有列均为空）
            if all(cell == "" for cell in row_values):
                continue
            row_dict = {
                headers[col]: self._parse_value(row_values[col])
                for col in range(len(headers)) if col < len(row_values)  # 防止列数不匹配
            }
            data.append(row_dict)

        return data

    def getSheetData(self, sheet_ref=None):
        """
        获取指定工作表的数据。
        Args:
            sheet_ref:
                - str: 工作表名称（如 "Sheet1"）
                - int: 工作表索引（从 0 开始）
                - list: 多个工作表名称或索引
                - None: 返回第一个工作表数据
        Returns:
            - 如果 sheet_ref 是 list，返回 {sheet_name: data} 的字典
            - 否则返回单个工作表的数据列表
        """
        workbook = self._load_workbook()
        try:
            # 处理多工作表请求
            if isinstance(sheet_ref, list):
                result = {}
                for ref in sheet_ref:
                    sheet = self._get_sheet(workbook, ref)
                    result[sheet.name] = self.readSheet(sheet)
                return result
            # 处理单工作表请求
            elif sheet_ref is not None:
                sheet = self._get_sheet(workbook, sheet_ref)
                return self.readSheet(sheet)
            else:
                # 默认返回第一个工作表
                sheet = workbook.sheet_by_index(0)
                return self.readSheet(sheet)
        finally:
            self._close_workbook()

    def _get_sheet(self, workbook, sheet_ref):
        """根据名称或索引返回工作表对象"""
        if isinstance(sheet_ref, (str)):
            try:
                return workbook.sheet_by_name(sheet_ref)
            except xlrd.XLRDError:
                raise ValueError("未找到工作表: {}".format(sheet_ref))
        elif isinstance(sheet_ref, int):
            try:
                return workbook.sheet_by_index(sheet_ref)
            except IndexError:
                raise IndexError("工作表索引越界: {}".format(sheet_ref))
        else:
            raise TypeError("sheet_ref 必须是 str 或 int")

    def getAllSheetData(self):
        """获取所有工作表的数据，返回 {sheet_name: data} 的字典"""
        workbook = self._load_workbook()
        try:
            return {
                sheet.name: self.readSheet(sheet)
                for sheet in workbook.sheets()
            }
        finally:
            self._close_workbook()

    def readAll(self):
        """兼容旧接口，等同于 getAllSheetData"""
        return self.getAllSheetData()
if __name__ == '__main__':
    path = r"C:\work\Project\Pre_support\Honor\PSI_script_0719\SIWAVE_TEST_CASE\SIWAVE_PDN_20240716.xlsx"
#     readSheet_Ironpython(path)
#     xls = XlsReader(path)
#     datas = xls.getAllSheetData()
#     print(datas)
